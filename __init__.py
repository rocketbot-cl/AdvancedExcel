# coding: utf-8

"""
Module to work with excel opened or created with rocketbot.

Rocketbot Functions:
    - GetParams("module"): Get the command name running. Module params in the package.json
    - GetParams("id"): Get the information sent by the user. Id params in the package.json
        var = GetParams(variable)
    - SetVar("variable_name", "dato"): Set a value to a variable.
    - GetVar("variable_name"): Get the value of a variable.
        var = GetVar("variable_name")

To install libraries use in the module path:
    pip install <package> -t ./libs 
"""

# Import globals or rocketbot libs
# -----------------------------------
# Changing the data types of all strings in the module at once
# from __future__ import unicode_literals
# from xlsx2csv import Xlsx2csv
import decimal
import io
import pandas as pd
from xlwings.constants import InsertShiftDirection
import xlwings as xw
import platform
import os
import sys
import subprocess
import dateutil.parser
import traceback
import json

# This lines is to linter
# -----------------------------------
GetParams = GetParams #type:ignore
tmp_global_obj = tmp_global_obj #type:ignore
PrintException = PrintException #type:ignore
SetVar = SetVar #type:ignore
GetGlobals = GetGlobals #type:ignore

# Add modules libraries to Rocektbot
# -----------------------------------
base_path = tmp_global_obj["basepath"]
cur_path = os.path.join(base_path, 'modules', 'AdvancedExcel', 'libs')

cur_path_x64 = os.path.join(cur_path, 'Windows' + os.sep +  'x64' + os.sep)
cur_path_x86 = os.path.join(cur_path, 'Windows' + os.sep +  'x86' + os.sep)

if sys.maxsize >= 2**32 and cur_path_x64 not in sys.path:
        sys.path.append(cur_path_x64)
if sys.maxsize < 2**32 and cur_path_x86 not in sys.path:
        sys.path.append(cur_path_x86)

def import_lib(relative_path, name, class_name=None):
    """
    - relative_path: library path from the module's libs folder
    - name: library name
    - class_name: class name to be imported. As 'from name import class_name'
    """

    import importlib.util

    cur_path = base_path + 'modules' + os.sep + \
        'AdvancedExcel' + os.sep + 'libs' + os.sep

    spec = importlib.util.spec_from_file_location(
        name, cur_path + relative_path)
    foo = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(foo)
    if class_name is not None:
        return getattr(foo, class_name)
    return foo


def get_date_with_format(xl_date, format_=None):
    import xlrd #type:ignore #ignore linter warnings
    datetime_date = xlrd.xldate_as_datetime(xl_date, 0)
    date_object = datetime_date.date()
    if format_ in [None, ""]:
        return date_object.isoformat()
    else:
        return datetime_date.strftime(format_)


def set_password(excel_file_path, pw):

    from pathlib import Path

    excel_file_path = Path(excel_file_path)

    vbs_script = \
    f"""' Save with password required upon opening

    Set excel_object = CreateObject("Excel.Application")
    Set workbook = excel_object.Workbooks.Open("{excel_file_path}")

    excel_object.DisplayAlerts = False
    excel_object.Visible = False

    workbook.SaveAs "{excel_file_path}",, "{pw}"

    excel_object.Application.Quit
    """

    # write
    vbs_script_path = excel_file_path.parent.joinpath("set_pw.vbs")
    with open(vbs_script_path, "w") as file:
        file.write(vbs_script)

    # execute
    subprocess.call(['cscript.exe', str(vbs_script_path)])

    # remove
    vbs_script_path.unlink()

    return None

module = GetParams("module")

# Get excel variables from Rocketbot
excel = GetGlobals("excel")
if excel.actual_id in excel.file_ and module != "Open":
    xls = excel.file_[excel.actual_id]
    wb = xls['workbook']

if module == "Open":
    id_ = GetParams("id")
    file_path = GetParams("path")
    password = GetParams("password")
    visible = GetParams("visible")
    var_ = GetParams("var_")
    try:

        app = xw.App(add_book=False)

        file_path = file_path.replace("/", os.sep)

        if platform.system() == "Windows":
            app.api.DisplayAlerts = False
            wb = app.api.Workbooks.Open(file_path, False, None, None, password, password, IgnoreReadOnlyRecommended=True, CorruptLoad=0)
            SetVar(var_, True)
        else:
            app.api.display_alerts = False
            wb = app.books.open(file_path, update_links = False, ignore_read_only_recommended = True)
            SetVar(var_, True)
        
        excel.actual_id = excel.id_default

        if id_:
            excel.actual_id = id_
        excel.file_[excel.actual_id] = {}
        excel.file_[excel.actual_id]['workbook'] = app.books[0]
        excel.file_[excel.actual_id]['app'] = excel.file_[
            excel.actual_id]['workbook'].app
        excel.file_[excel.actual_id]['sheet'] = excel.file_[
            excel.actual_id]['workbook'].sheets[0]
        excel.file_[excel.actual_id]['path'] = file_path

    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        SetVar(var_, False)
        raise e

if module == "Calculate":
    calculation_type = GetParams("calculation_type")
    calculate = GetParams("calculate")
    
    try:
        if calculation_type:
            wb.api.Application.Calculation = eval(calculation_type)
        
        if calculate and eval(calculate):
            wb.api.Application.Calculate()
        
    except Exception as e:
        traceback.print_exc()
        PrintException()
        raise e

if module == "ReadCells":
    range_ = GetParams("range")
    sheet_ = GetParams("sheet")
    var_ = GetParams("var_")
    date_format = GetParams("date_format")
    custom = GetParams("custom")

    try:
        if not sheet_:
            sheet = wb.sheets.active
        else:
            wb_sheets = [sh.name for sh in wb.sheets]
            sheet=None
            for s in wb_sheets:
                if s.strip() == sheet_:
                    sheet = wb.sheets(s)
                    break
            if not sheet:
                raise Exception(f"The name {sheet_} does not exist in the book")
            
        if date_format == "date1":
            date_format = '%d-%m-%Y'
        elif date_format == "date2":
            date_format = '%d-%m-%y'
        elif date_format == "date3":
            date_format = '%Y-%m-%d'
        elif date_format == "date4":
            date_format = '%m-%d-%Y'
        elif date_format == 'custom':
            date_format = custom
        else:
            date_format = None

        global _values    
        _values = sheet.api.Range(range_).Value2
        print(_values)
        value = None
        if date_format is not None:
            valueGotten = xw.sheets[sheet].range(range_).value
            cont = 0
            info = []
            try:
                for each in valueGotten:
                    cont += 1
            except:
                cont = 1
            if (cont > 1):
                for each in valueGotten:
                    value_date = each.strftime(date_format)
                    info.append(value_date)
            else:
                valueGotten = valueGotten.strftime(date_format)
                info.append(valueGotten)
            print(info)
            SetVar(var_, info)
            
        if date_format is None:
            if isinstance(_values, tuple):
                value = [list(i) for i in _values if isinstance(_values, tuple)]
            else:
                value = _values
        
            value = value if value != [] else _values
            SetVar(var_, value)
        
    except Exception as e:
        traceback.print_exc()
        PrintException()
        raise e

if module == "ExcelDateToDate":
    date_ = GetParams("excel_date")
    format_ = GetParams("format")
    var_ = GetParams("var_")
    
    try:
        value = get_date_with_format(int(date_), format_)
        
        SetVar(var_, value)
    except Exception as e:
        traceback.print_exc()
        PrintException()
        raise e

if module == "CellColor":

    import traceback
    
    range_ = GetParams("range")
    sheet_ = GetParams("sheet")
    all_ = GetParams("all")
    color = GetParams("color")
    custom = GetParams("custom")
    
    try:
        if color == "red":
            rgb = (255, 0, 0)

        elif color == "blue":
            rgb = (0, 0, 255)
        elif color == "green":
            rgb = (0, 255, 0)
        elif color == "grey":
            rgb = (130, 130, 130)
        elif color == "yell":
            rgb = (255, 255, 0)
        else:
            rgb = eval(custom)
        
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break       
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")        
                
        if all_ and eval(all_) == True:
            xw.sheets[sheet].api.Cells.Interior.Color = xw.utils.rgb_to_int(rgb)
        else:            
            if sheet_:
                xw.sheets[sheet].range(range_).color = rgb
            elif range_:
                xw.Range(range_).color = rgb
            else:
                raise Exception('Must select sheet or range.')

    except Exception as e:
        traceback.print_exc()
        PrintException()
        raise e

if module == "GetColor":
    
    sheet_ = GetParams("sheet")
    cell = GetParams("cell")
    res = GetParams("res")
    
    colors = []
    
    wb_sheets = [sh.name for sh in wb.sheets]
    sheet=None
    for s in wb_sheets:
        if s.strip() == sheet_:
            sheet = s
            break
    if not sheet:
        raise Exception(f"The name {sheet_} does not exist in the book") 
    try:
        background = xw.sheets[sheet].range(cell).color
        # Range object in xlwings has no font property, so I use the API (py32win) wich returns an int
        font = xw.sheets[sheet].range(cell).api.Font.Color
        # Int to RGB
        font_ = xw.utils.int_to_rgb(font)
        colors.append(background)
        colors.append(font_)
        SetVar(res, colors)
    
    except Exception as e:
        PrintException()
        raise e

if module == "GetCellFormats":
    
    sheet_ = GetParams("sheet")
    cell = GetParams("cell")
    res = GetParams("res")
    
    try:
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")    
        cell = xw.sheets[sheet].range(cell)
        
        horizontal_alignment_glossary = {
            -4108: "xlHAlignCenter",
            -4131: "xlHAlignLeft",
            -4152: "xlHAlignRight",
            -4117: "xlHAlignDistributed",
            1: "xlHAlignGeneral",
            -4130: "xlHAlignJustify",
            7: "xlHAlignCenterAcrossSelection",
            5: "xlHAlignFill"
        }

        vertical_alignment_glossary = {
            -4107: "xlVAlignBottom",
            -4108: "xlVAlignCenter",
            -4117: "xlVAlignDistributed",
            -4160: "xlVAlignTop",
            -4130: "xlVAlignJustify",
        }

        font_underline_glossary = {
            -4142: "xlUnderlineStyleNone",
            2: "xlUnderlineStyleSingle",
            -4119: "xlUnderlineStyleDouble",
            4: "xlUnderlineStyleSingleAccounting",
            5: "xlUnderlineStyleDoubleAccounting",
        }
        
        formats = {
            'FontColor': xw.utils.int_to_rgb(cell.api.Font.Color),
            'BackgroundColor': xw.utils.int_to_rgb(cell.api.Interior.Color),
            'HorizontalAlignment': horizontal_alignment_glossary.get(cell.api.HorizontalAlignment),
            'VerticalAlignment': vertical_alignment_glossary.get(cell.api.VerticalAlignment),
            'Indentation': cell.api.IndentLevel,
            'FontName': cell.api.Font.Name,
            'FontSize': cell.api.Font.Size,
            'FontBold': cell.api.Font.Bold,
            'FontItalic': cell.api.Font.Italic,
            'FontUnderline': font_underline_glossary.get(cell.api.Font.Underline),
            'FontStrikethrough': cell.api.Font.Strikethrough,
            'FontSubscript': cell.api.Font.Subscript,
            'FontSuperscript': cell.api.Font.Superscript,
            'NumberFormat': cell.api.NumberFormat,
            'CellWidth': cell.api.ColumnWidth,
            'CellHeight': cell.api.RowHeight,
            'WrapText': cell.api.WrapText,
            'MergeCells': cell.api.MergeCells,
            'Locked': cell.api.Locked
        }
        
        
        SetVar(res, formats)
    
    except Exception as e:
        PrintException()
        raise e

if module == "InsertFormula":
    
    sheet_ = GetParams("sheet")
    cell = GetParams("cell")
    formula = GetParams("formula")
    no_iie = GetParams("no_iie")
    
    if not sheet_:
        sheet = wb.sheets.active
    else:
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = wb.sheets(s)
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
           
        
    if no_iie and eval(no_iie):
        sheet.range(cell).api.Formula2 = formula
    else:
        sheet.range(cell).formula = formula

if module == "InsertMacro":
    macro = GetParams("macro_path")

    content_macro = None
    if macro and macro != "ERROR_NOT_VAR":
        if os.path.exists(macro):
            with open(macro, "r", encoding="latin-1") as m:
                content_macro = m.read()
                m.close()
        else:
            raise Exception("No existe el archivo de macro")
    else:
        raise Exception("No existe variable con ruta de macro")
    if content_macro is not None:
        tmp = xls['workbook'].api.VBProject.VBComponents.Add(1)
        tmp.CodeModule.AddFromString(content_macro.strip())

if module == "SelectCells":

    cells = GetParams("cells")
    copy = GetParams("copy")
    sheet_ = GetParams("sheet_name")

    if copy is None:
        copy = False

    try:
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        try:
            wb.sheets[sheet].select()
        except:
            pass
        try:
            wb.sheets[sheet].range(cells).select()
        except:
            pass
                
        if copy:
            wb.sheets[sheet].range(cells).options(ndim=2).copy()
            
    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "CutCopyMode":

    try:
        wb.api.Application.CutCopyMode = False

    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if (module == "getCurrencyValue"):

    sheet_ = GetParams("sheetWanted")
    cellRange = GetParams("cellRange")
    finalResult = []
    
    wb_sheets = [sh.name for sh in wb.sheets]
    sheet=None
    for s in wb_sheets:
        if s.strip() == sheet_:
            sheet = s
            break
    if not sheet:
        raise Exception(f"The name {sheet_} does not exist in the book")
    
    valueGotten = xw.sheets[sheet].range(cellRange).value
    cont = 1
    try:
        if isinstance(valueGotten, list):
            cont = len(valueGotten)
    except:
        cont = 1

    if (cont > 1):
        for each in valueGotten:
            if isinstance(each, list):
                list_ = []
                for each2 in each:
                    try:
                        list_.append(float(each2))
                    except:
                        list_.append(each2)
                finalResult.append(list_)
            else:
                try:
                    finalResult.append(float(each))
                except:
                    finalResult.append(each)
    else:
        try:
            finalResult.append(float(valueGotten))
        except:
            finalResult.append(valueGotten)

    whereToStoreData = GetParams("whereToStoreData")
    SetVar(whereToStoreData, finalResult)


if (module == "getDateValue"):

    sheet_ = GetParams("sheetWanted")
    cellRange = GetParams("cellRange")
    finalResult = []
    
    wb_sheets = [sh.name for sh in wb.sheets]
    sheet=None
    for s in wb_sheets:
        if s.strip() == sheet_:
            sheet = s
            break
    if not sheet:
        raise Exception(f"The name {sheet_} does not exist in the book")
    
    valueGotten = xw.sheets[sheet].range(cellRange).value
    cont = 0
    try:
        try:
            for each in valueGotten:
                cont += 1
        except:
            cont = 1

        if (cont > 1):
            for each in valueGotten:
                value_date = each.strftime("%d/%m/%Y %H:%M:%S")
                finalResult.append(value_date)
        else:
            valueGotten = valueGotten.strftime("%d/%m/%Y %H:%M:%S")
            finalResult.append(valueGotten)
        whereToStoreData = GetParams("whereToStoreData")
        SetVar(whereToStoreData, finalResult)

    except Exception as e:
        PrintException()
        raise e

if module == "copyPaste":
    rango1 = GetParams("cell_range1")
    rango2 = GetParams("cell_range2")
    hoja1_ = GetParams("sheet_name1")
    hoja2_ = GetParams("sheet_name2")
    option = GetParams("option")
    ope = GetParams("operation")
    saltar = GetParams("skip_blanks")
    trans = GetParams("transpose")
    
    try:
        args = {}
        
        if option:
            args['paste'] = option
        
        if ope:
            args['operation'] = ope
        
        if saltar and saltar != "False":
            args['skip_blanks'] = True
        
        if trans and trans != "False":
            args['transpose'] = True
        
        wb_sheets = [sh.name for sh in wb.sheets]
        hoja1=None
        hoja2=None
        for s in wb_sheets:
            if s.strip() == hoja1_:
                hoja1 = s
                break
        for s in wb_sheets:
            if s.strip() == hoja2_:
                hoja2 = s
                break
        
        if not hoja1:
            raise Exception(f"The name {hoja1_} does not exist in the book")
        if not hoja2:
            raise Exception(f"The name {hoja2_} does not exist in the book")
        
        wb.sheets[hoja1].range(rango1).options(ndim=2).copy()
        wb.sheets[hoja2].range(rango2).paste(**args)
    
    except Exception as e:
        PrintException()
        raise e
    
if module == "formatCell":
    sheet_ = GetParams("sheet_name")
    rango = GetParams("cell_range")
    formato = GetParams("format_")
    custom = GetParams("custom")
    texttoval = GetParams("texttoval")
    
    import re 
    try:
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")

        if len(rango) == 1:
            rango = rango + ':' + rango
        
        if formato == "text":
            wb.sheets[sheet].range(rango).number_format = '@'

        def transform_data(cell=None, row=None):
            global dateutil
            new_row = []
            if row:
                for cell in row:
                    try:
                        if cell.strip().isnumeric():
                            cell = float(cell)
                            new_range.append(cell)
                        elif "/" in cell or "-" in cell:
                            cell_date = dateutil.parser.parse(cell, dayfirst=True)
                            new_row.append(cell_date)
                        else:
                            new_range.append(float(cell.replace(".","").replace(",", ".")))
                    except Exception as e:
                        new_row.append(cell) 
                new_range.append(new_row)
            else:
                try:
                    if cell.strip().isnumeric():
                        cell = float(cell)
                        new_range.append(cell)
                    elif "/" in cell or "-" in cell:
                        cell_date = dateutil.parser.parse(cell, dayfirst=True)
                        new_range.append(cell_date)
                    else:
                        new_range.append(float(cell.replace(".","").replace(",", ".")))
                except Exception as e:
                    new_range.append(cell)

        if texttoval and eval(texttoval) == True:
            global new_range
            new_range = []
            # Multiple rows
            if isinstance(wb.sheets[sheet].range(rango).value[0], list):
                for row in wb.sheets[sheet].range(rango).value:
                    transform_data(row=row)
                    
            # Single row
            elif isinstance(wb.sheets[sheet].range(rango).value, list):
                for cell in wb.sheets[sheet].range(rango).value:
                    transform_data(cell=cell)
                    
            # Cell
            else:
                cell = wb.sheets[sheet].range(rango).value
                transform_data(cell=cell)
                            
            # This block of code is used when the 'Text o Value' is applied to one column. Where the range.value returns a sigle list.
            # So, when re-writing it has to be refactore to a list of lists (with one element each).
            # The Try/except is used to avoid errors when the ranges is a single cell.
            try:
                regex = "([a-zA-Z]+)[0-9]?:([a-zA-Z]?)[0-9]?"
                matches = re.match(regex, rango).groups()
                if matches[0] == matches[1]:
                    new_range = [[i] for i in new_range]
            except:
                pass
                
            wb.sheets[sheet].range(rango).value = new_range
        
        if formato == "number_":
            numbers = wb.sheets[sheet].range(rango).value
            d = 0
            if type(numbers[0]) != list and len(numbers) == 1:
                numbers = [numbers]

            for i in range(len(numbers)):
                element = numbers[i]
                if type(element) == list:
                    for idx in range(len(element)):

                        number = element[idx]

                        if type(element[idx]) is str:
                            number = number.split(",")
                            if "." in number[0]:
                                number[0] = number[0].replace(".", "")
                            number = ".".join(number)
                            if d < len(str(number).split(".")[1]):
                                d = len(str(number).split(".")[1])
                        element[idx] = number
                        numbers[i] = element
                else:
                    number = numbers[i]
                    if type(number) is str:
                        number = number.split(",")
                        if "." in number[0]:
                            number[0] = number[0].replace(".", "")
                        number = ".".join(number)
                        tmp = str(number).split(".")
                        if len(tmp) > 1:
                            if d < len(tmp[1]):
                                d = len(tmp[1])
                    numbers[i] = number

            if rango.split(":")[0][0] == rango.split(":")[1][0]:
                for i in range(len(numbers)):
                    numbers[i] = [numbers[i]]

            wb.sheets[sheet].range(rango).value = numbers
            print("format", wb.sheets[sheet].range(rango).number_format)
            if d == 0:
                wb.sheets[sheet].range(rango).number_format = '0'
            else:
                wb.sheets[sheet].range(rango).number_format = '0,{}'.format('0' * d)
        
        if formato == "coin_":
            wb.sheets[sheet].range(rango).number_format = '$#.##0'

        if formato == "date1":
            wb.sheets[sheet].range(rango).number_format = 'dd-mm-yyyy'

        if formato == "date2":
            wb.sheets[sheet].range(rango).number_format = 'dd-mm-yy'

        if formato == "date3":
            wb.sheets[sheet].range(rango).number_format = 'yyyy-mm-dd'

        if formato == "decimal1":
            wb.sheets[sheet].range(rango).number_format = '0,0'

        if formato == "decimal2":
            wb.sheets[sheet].range(rango).number_format = '#.##0,0'

        if formato == "long_date":
            wb.sheets[sheet].range(rango).number_format = 'dd/mm/yyyy h:mm:ss'
        
        if formato == 'custom':
            wb.sheets[sheet].range(rango).number_format = custom        
            
    except Exception as e:
        PrintException()
        traceback.print_exc()
        raise e


if module == "clearCell":
    sheet_ = GetParams("sheet")
    range_ = GetParams("cell_range")
    
    wb_sheets = [sh.name for sh in wb.sheets]
    sheet=None
    for s in wb_sheets:
        if s.strip() == sheet_:
            sheet = s
            break
    if not sheet:
        raise Exception(f"The name {sheet_} does not exist in the book")
    
    try:
        sheet_selected = wb.sheets[sheet]
        
        sheet_selected.api.Range(range_).ClearContents()
        
    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "createSheet":
    hoja = GetParams("sheet_name")
    last = GetParams("after")

    try:

        if not last:
            res = [a.name for a in wb.sheets]
            last = res[-1]

        wb.sheets.add(name=hoja, after=last)
    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "deleteSheet":

    sheet_ = GetParams("sheet_name")
    var_ = GetParams("var_")
    res = False

    wb_sheets = [sh.name for sh in wb.sheets]
    sheet=None
    for s in wb_sheets:
        if s.strip() == sheet_:
            sheet = s
            break
    if not sheet:
        SetVar(var_, res)
        raise Exception(f"The name {sheet_} does not exist in the book")
    
    wb.sheets[sheet].delete()
    res = True
    SetVar(var_, res)

if module == "copy_other":

    try:
        excel1 = GetParams("excel1")
        id_ = GetParams("id")
        excel2 = GetParams("excel2")
        hoja1_ = GetParams("sheet_name1")
        hoja2_ = GetParams("sheet_name2")
        rango1 = GetParams("cell_range1")
        rango2 = GetParams("cell_range2")
        only_values = GetParams("values")
        normal = GetParams("normal")
        password = GetParams("password")              
        
        load = 0 if normal and eval(normal) else 2

        if only_values is not None:
            only_values = eval(only_values)

        global excel__
        if platform.system() == "Windows":
            if excel1:
                open = 0
                app = None
                excel__ = excel
                paths = [excel__.file_[i]['path'].replace("\\","/") for i in list(excel__.file_.keys())]
                if excel1.replace("\\","/") not in paths:
                    try:
                        wb1 = wb.app.books.api.Open(excel1, False, None, None, password, password, IgnoreReadOnlyRecommended=True, CorruptLoad=load)
                    except:
                        app = xw.App(add_book=False)
                        app.api.DisplayAlerts = False
                        app.api.Visible = False
                        wb1 = app.api.Workbooks.Open(excel1, False, None, None, password, password, IgnoreReadOnlyRecommended=True, CorruptLoad=load)
                        wb = app.books[0]
                    open = 1
                else:
                    try:
                        wb1 = wb.api
                        wb1.Sheets
                    except:
                        try:
                            wb1 = wb.app.books.api.Open(excel1, False, None, None, password, password, IgnoreReadOnlyRecommended=True, CorruptLoad=load)
                        except:
                            app = xw.App(add_book=False)
                            app.api.DisplayAlerts = False
                            app.api.Visible = False
                            wb1 = app.api.Workbooks.Open(excel1, False, None, None, password, password, IgnoreReadOnlyRecommended=True, CorruptLoad=load)
                            wb = app.books[0]
                        open = 1
            elif id_:
                if id_ not in excel.file_:
                    raise Exception(f"The id {id_} does not exist.")
                xls = excel.file_[id_]
                wb1 = xls['workbook'].api
            else:
                wb1 = wb.api

            wb_sheets1 = [sh.Name for sh in wb1.Sheets]
            hoja1 = None
            for s in wb_sheets1:
                if s.strip() == hoja1_:
                    hoja1 = s
                    break
            if not hoja1:
                raise Exception(f"The name {hoja1_} does not exist in the book")
            
            origin_sheet = wb1.Sheets(hoja1)
            my_values = origin_sheet.Range(rango1)
            
            wb2 = wb.app.books.api.Open(excel2, False, None, None, password, password, IgnoreReadOnlyRecommended=True, CorruptLoad=load)
            
            wb_sheets2 = [sh.Name for sh in wb2.Sheets]
            hoja2 = None
            for s in wb_sheets2:
                if s.strip() == hoja2_:
                    hoja2 = s
                    break
            if not hoja2:
                raise Exception(f"The name {hoja2_} does not exist in the book")
            
            destiny_sheet = wb2.Sheets(hoja2)
            
            if not rango2:
                rango2 = "A" + str(destiny_sheet.UsedRange.Rows.Count)
            if len(rango2) == 1 and not rango2[0].isnumeric():
                rango2 = rango2 + str(destiny_sheet.UsedRange.Rows.Count)

            if not only_values:
                origin_sheet.Range(rango1).Copy(
                    destiny_sheet.Range(rango2))
            else:
                destiny_sheet.Range(rango2).Value = my_values.Value
                
            wb2.Application.DisplayAlerts = False
            wb2.SaveAs(excel2.replace("/",os.sep))
            wb2.Close()
            if open == 1:
                wb1.Close()
                if app is not None:
                    app.kill()
        else:
            if excel1:
                open = 0
                app = None
                excel__ = excel
                paths = [excel__.file_[i]['path'].replace("\\","/") for i in list(excel__.file_.keys())]
                if excel1.replace("\\","/") not in paths:
                    try:
                        wb1 = wb.app.books.open(excel1, update_links = False, ignore_read_only_recommended = True)
                    except:
                        app = xw.App(add_book=False)
                        app.api.display_alerts = False
                        app.api.visible = False
                        wb1 = app.books.open(excel1, update_links = False, ignore_read_only_recommended = True)
                        wb = app.books[0]
                    open = 1
                else:
                    try:
                        wb1 = wb
                    except:
                        try:
                            wb1 = wb.app.books.open(excel1, update_links = False, ignore_read_only_recommended = True)
                        except:
                            app = xw.App(add_book=False)
                            app.api.display_alerts = False
                            app.api.visible = False
                            wb1 = app.books.open(excel1, update_links = False, ignore_read_only_recommended = True)
                            wb = app.books[0]
                        open = 1
            elif id_:
                if id_ not in excel.file_:
                    raise Exception(f"The id {id_} does not exist.")
                xls = excel.file_[id_]
                wb1 = xls['workbook']
            else:
                wb1 = wb
                
            wb_sheets1 = [sh.name for sh in wb1.sheets]
            for s in wb_sheets1:
                if s.strip() == hoja1_:
                    hoja1 = s
                    break
            if not hoja1:
                raise Exception(f"The name {hoja1_} does not exist in the book")

            origin_sheet = wb1.sheets[hoja1]
            my_values = origin_sheet.range(rango1)
            
            values = my_values.value
            wb2 = wb.app.books.open(excel2, update_links = False, ignore_read_only_recommended = True)
            
            wb_sheets2 = [sh.name for sh in wb2.sheets]
            for s in wb_sheets2:
                if s.strip() == hoja2_:
                    hoja2 = s
                    break
            if not hoja2:
                raise Exception(f"The name {hoja2_} does not exist in the book")
            
            destiny_sheet = wb2.sheets(hoja2)
            if not only_values:
                origin_sheet.range(rango1).copy(
                    destiny_sheet.range(rango2))
            else:
                destiny_sheet.range(rango2).value = values
                
            wb2.app.api.display_alerts = False
            wb2.api.save_as(excel2.replace("/",os.sep))
            wb2.close()
            if open == 1:
                wb1.close()
                if app is not None:
                    app.kill()

    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "addRow":

    try:
        sheet_ = GetParams("sheet")
        row = GetParams("row_")
        tipo = GetParams("type_")
        opcion_ = GetParams("option_")

        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")

        sheet_selected = wb.sheets[sheet]

        try:
            import re
            regex = "([a-zA-Z]*)([0-9]*)[:]?([a-zA-Z]*)([0-9]*)"
            matches = re.match(regex, row).groups()
        except:
            pass
        
        if opcion_ == "add_":

            if platform.system() == 'Windows':
                if ":" in row:
                    if tipo == "down_":
                        if matches[0] and matches[2]:
                            fila = matches[0] + str(int(matches[1])+1) \
                                + ":" + matches[2] + str(int(matches[3])+1)
                        else:
                            fila = str(int(matches[1])+1) + ":" + str(int(matches[3])+1)
                            
                        sheet_selected.range(fila).api.Insert(
                            InsertShiftDirection.xlShiftDown)

                    if tipo == "up_":
                        sheet_selected.range(row).api.Insert(
                            InsertShiftDirection.xlShiftDown)

                else:
                    if tipo == "down_":
                        if matches[0]:
                            fila = matches[0] + str(int(matches[1])+1)
                        else:
                            fila = str(int(matches[1])+1) + ":" + str(int(matches[1])+1)

                        sheet_selected.range(fila).api.Insert(
                            InsertShiftDirection.xlShiftDown)

                    if tipo == "up_":
                        if matches[0]:
                            fila = row
                        else:
                            fila = str(matches[1]) + ":" + str(matches[1])
                            
                        sheet_selected.range(fila).api.Insert(
                            InsertShiftDirection.xlShiftDown)

            else:
                if ":" in row:
                    if tipo == "down_":
                        if matches[0] and matches[2]:
                            fila = matches[0] + str(int(matches[1])+1) \
                                + ":" + matches[2] + str(int(matches[3])+1)
                        else:
                            fila = str(matches[0]+1) + ":" + str(matches[2]+1)
                            
                        sheet_selected.api.Rows[fila].insert_into_range()
                    if tipo == "up_":
                        sheet_selected.api.Rows[row].insert_into_range()

                else:
                    if tipo == "down_":
                        if matches[0]:
                            fila = matches[0] + str(int(matches[1])+1)
                        else:
                            fila = str(matches[1]+1) + ":" + str(matches[1]+1)

                        sheet_selected.api.Rows[fila].insert_into_range()

                    if tipo == "up_":
                        if matches[0]:
                            fila = row
                        else:
                            fila = str(matches[1]) + ":" + str(matches[1])
                            
                        sheet_selected.api.Rows[fila].insert_into_range()

        if opcion_ == "delete_":
            
            if ":" in row:
                fila = row
            elif ":" not in row and matches[0]:
                fila = row
            else:
                fila = str(matches[1]) + ":" + str(matches[1])
            
            sheet_selected.range(fila).api.Delete()

    except Exception as e:
        PrintException()
        raise e

if module == "addCol":
    try:
        sheet_ = GetParams("sheet")
        col_ = GetParams("col_")
        opcion_ = GetParams("option_")
        
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        
        try:
            import re
            regex = "([a-zA-Z]*)([0-9]*)[:]?([a-zA-Z]*)([0-9]*)"
            matches = re.match(regex, col_).groups()
        except:
            pass

        if opcion_ == "add_":

            if platform.system() == 'Windows':

                if ":" in col_:
                    wb.sheets[sheet].range(col_).api.Insert(
                        InsertShiftDirection.xlShiftToRight)
                else:
                    if matches[0] and matches[1]:
                        col = col_
                    else:
                        col = str(matches[0]) + ":" + str(matches[0])
                    
                    wb.sheets[sheet].range(col).api.Insert(
                        InsertShiftDirection.xlShiftToRight)

            else:
                if ":" in col_:
                    wb.sheets[sheet].api.columns[col_].insert_into_range()
                else:
                    if matches[0] and matches[1]:
                        col = col_
                    else:
                        col = str(matches[0]) + ":" + str(matches[0])
                    wb.sheets[sheet].api.columns[col].insert_into_range()

        if opcion_ == "delete_":
            if platform.system() == 'Windows':
                if ":" in col_:
                    wb.sheets[sheet].range(col_).api.Delete()
                else:
                    if matches[0] and matches[1]:
                        col = col_
                    else:
                        col = str(matches[0]) + ":" + str(matches[0])
                    wb.sheets[sheet].range(col).api.Delete()

            else:
                if ":" in col_:
                    wb.sheets[sheet].range(col_).api.delete()
                else:
                    if matches[0] and matches[1]:
                        col = col_
                    else:
                        col = str(matches[0]) + ":" + str(matches[0])
                    wb.sheets[sheet].range(col).api.delete()

    except Exception as e:
        PrintException()
        raise e

if module == "csvToxlsx":
    csv_path = GetParams("csv_path")
    xlsx_path = GetParams("xlsx_path")
    sep = GetParams("separator") or ","
    with_header = GetParams("header")
    encoding = GetParams("encoding")
    
    import string
    global printable
    printable = set(string.printable)
    
    if not encoding or encoding == '':
        encoding = 'latin-1'
    
    try:
        if not csv_path or not xlsx_path:
            raise Exception("Falta una ruta")

        import csv
        from openpyxl import Workbook, load_workbook

        if platform.system() == "Windows":
            import ctypes as ct

            csv.field_size_limit(int(ct.c_ulong(-1).value // 2))
            limit1 = csv.field_size_limit()
        if sep.startswith("\\t"):
            sep = "\t"
        workbook = Workbook()
        workbook.encoding = encoding
        worksheet = workbook.active
        with open(csv_path, "r", encoding=encoding) as fobj:
            csv_reader = csv.reader((x.replace('\0', '') for x in fobj), delimiter=sep)
            for row_index, row in enumerate(csv_reader):
                for col_index, value in enumerate(row):
                    try:
                        value = value.decode()
                    except:
                        pass
                    try:                      
                        worksheet.cell(row_index + 1, col_index + 1).value = value
                    except:
                        # Eliminates non printeable characters to avoid writing errors
                        value = ''.join(filter(lambda x: x in printable, value))
                        worksheet.cell(row_index + 1, col_index + 1).value = value  

        workbook.save(xlsx_path)

    except Exception as e:
        print("\x1B[" + "31;40mAn error occurred\x1B[" + "0m")
        PrintException()
        raise e

if module == "xlsxToCsv":
    csv_path = GetParams("csv_path")
    xlsx_path = GetParams("xlsx_path")
    delimiter = GetParams("delimiter")
    sheet_name = GetParams("sheet_name")
    import csv
    from openpyxl import Workbook, load_workbook
    
    try:
        if delimiter == "\\t":
            delimiter = "\t"
        if not delimiter:
            delimiter = ","

        if not sheet_name:
            sheet_name = "Sheet0"
        data_xls = load_workbook(xlsx_path)[sheet_name]
        data = [[str(data).replace("\xa0", "") for data in row]
                for row in data_xls.iter_rows(values_only=True)]
        # data_xls = pd.read_excel(xlsx_path, sheet_name, index_col=None, header=None)

        with open(csv_path, mode='w', newline='') as csv_file:
            csv_writer = csv.writer(
                csv_file, delimiter=delimiter, quotechar='"', quoting=csv.QUOTE_MINIMAL)
            for row in data:
                csv_writer.writerow(row)

        # data_xls.to_csv(csv_path, encoding='utf-8', index=False, header=False)
        # Xlsx2csv(xlsx_path, outputencoding="utf-8", delimiter=delimiter, floatformat=True).convert(csv_path)
    except Exception as e:
        traceback.print_exc()
        PrintException()
        raise e

if module == "xlsx_to_csv":
    csv_path = GetParams("csv_path")
    xlsx_path = GetParams("xlsx_path")
    delimiter = GetParams("delimiter")
    sheet_name = GetParams("sheet_name")

    try:
        import xlwings as xw
        import csv

        app = xw.App(visible=False)
        wb = xw.Book(xlsx_path)
        sheet = wb.sheets[sheet_name]
        
        # The VBA API was dicarded because it has trouble with special characters
        # wb.sheets[sheet_name].api.Copy()
        # xw.books.active.api.SaveAs(csv_path, 6, ConflictResolution = 2)
        
        sheet_ = []
        used_range = sheet.used_range.value
        try:
            for row in used_range:
                row_ = []
                for value in row:
                    if isinstance(value, str):
                        value = ''.join(i for i in value if ord(i)<128)
                    if isinstance(value, float):
                        if value.is_integer():
                            value = int(value)
                    row_.append(value)
                sheet_.append(row_)
        # When the sheet to convert has one row.
        except TypeError:
            for value in used_range:
                if isinstance(value, str):
                    value = ''.join(i for i in value if ord(i)<128)
                if isinstance(value, float):
                    if value.is_integer():
                        value = int(value)
                row_.append(value)
            sheet_.append(row_)
                 
        with open(csv_path, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile, delimiter=delimiter, quotechar='|', quoting=csv.QUOTE_MINIMAL)
            writer.writerows(sheet_)

        wb.close()
        app.kill()
        
    except Exception as e:
        wb.close()
        app.kill()
        PrintException()
        raise e

if module == "countColumns":
    sheet_ = GetParams("sheet")
    column_name = GetParams("column")
    result = GetParams("var_")

    try:
        
        if not sheet_:
            sheet = wb.sheets.active.name
        else:
            wb_sheets = [sh.name for sh in wb.sheets]
            sheet=None
            for s in wb_sheets:
                if s.strip() == sheet_:
                    sheet = s
                    break
            if not sheet:
                raise Exception(f"The name {sheet_} does not exist in the book")
            
        sheet_selected = wb.sheets[sheet]
            
        col = sheet_selected.used_range.last_cell.column
        
        if column_name and eval(column_name):
            col = wb.sheets[sheet].cells(1,col).get_address()

        if result:
            SetVar(result, col)

    except Exception as e:
        PrintException()
        raise e

if module == "countRows":

    sheet_ = GetParams("sheet")
    row_ = GetParams("row_") #Column
    result = GetParams("var_")
    countAll = GetParams("countAll")
    if countAll is not None:
            countAll = eval(countAll)

    if not sheet_:
        sheet = wb.sheets.active.name
    else:
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
    
    if not row_:
        row_ = 'A'

    try:
        
        if countAll == True:
            total = wb.sheets[sheet].api.UsedRange.Rows.Count

        else:
            # Selects the whole row, goes to the last cell and goes up until find the first cell with value.
            a = wb.sheets[sheet].range(f"{row_}:{row_}").last_cell
            while not a.value:
                a = a.end('up')
                if a.row == 1 and not a.value:
                    break

            total = wb.sheets[sheet].range(row_ + str(a.row)).row
        if result:
            SetVar(result, total)

    except Exception as e:
        PrintException()
        raise e

if module == "hide":
    sheet_ = GetParams("sheet")
    range = GetParams("range")
    result = GetParams("var_")

    try:
        if not sheet_:
            sheet = wb.sheets.active.name
        else:
            wb_sheets = [sh.name for sh in wb.sheets]
            sheet=None
            for s in wb_sheets:
                if s.strip() == sheet_:
                    sheet = s
                    break
            if not sheet:
                raise Exception(f"The name {sheet_} does not exist in the book")

        sheet_selected = wb.sheets[sheet]

        if range.replace(':', '').isnumeric():
            if ":" in range:
                sheet_selected.range(range).api.EntireRow.Hidden = True
            else:
                range_ = range + ":" + range
                sheet_selected.range(range_).api.EntireRow.Hidden = True
            if result:
                SetVar(result, True)
        elif all(c.isalpha() or c == ':' for c in range):
            if ":" in range:
                sheet_selected.range(range).api.EntireColumn.Hidden = True
            else:
                range_ = range + ":" + range
                sheet_selected.range(range_).api.EntireColumn.Hidden = True
            if result:
                SetVar(result, True)
        else:
            if result:
                SetVar(result, False)
            raise Exception(f"El rango {range} no es valido.")    
    except Exception as e:
        PrintException()
        raise e

if module == "show":
    sheet_ = GetParams("sheet")
    range = GetParams("range")
    result = GetParams("var_")

    try:
        if not sheet_:
            sheet = wb.sheets.active.name
        else:
            wb_sheets = [sh.name for sh in wb.sheets]
            sheet=None
            for s in wb_sheets:
                if s.strip() == sheet_:
                    sheet = s
                    break
            if not sheet:
                raise Exception(f"The name {sheet_} does not exist in the book")
            
        sheet_selected = wb.sheets[sheet]
            
        if range.replace(':', '').isnumeric():
            if ":" in range:
                sheet_selected.range(range).api.EntireRow.Hidden = False
            else:
                range_ = range + ":" + range
                sheet_selected.range(range_).api.EntireRow.Hidden = False
            if result:
                SetVar(result, True)
        elif all(c.isalpha() or c == ':' for c in range):
            if ":" in range:
                sheet_selected.range(range).api.EntireColumn.Hidden = False
            else:
                range_ = range + ":" + range
                sheet_selected.range(range_).api.EntireColumn.Hidden = False
            if result:
                SetVar(result, True)
        else:
            if result:
                SetVar(result, False)
            raise Exception(f"El rango {range} no es valido.")    
    except Exception as e:
        PrintException()
        raise e
    
if module == "xlsToxlsx":

    xls_path = GetParams('xls_path')
    xlsx_path = GetParams('xlsx_path')

    try:
        
        if sys.maxsize > 2**32:
            p = import_lib(f"Windows{os.sep}x64{os.sep}pyexcel{os.sep}__init__.py", "pyexcel") # import pyexcel as p
        if sys.maxsize > 32:
            p = import_lib(f"Windows{os.sep}x86{os.sep}pyexcel{os.sep}__init__.py", "pyexcel") # import pyexcel as p
        
        try:
            
            p.save_book_as(file_name=xls_path,
                           dest_file_name=xlsx_path)
        except:
            if sys.maxsize > 2**32:
                Workbook = import_lib(f"Windows{os.sep}x64{os.sep}xlwt{os.sep}__init__.py", "xlwt", "Workbook") # from xlwt import Workbook
            if sys.maxsize > 32:
                Workbook = import_lib(f"Windows{os.sep}x86{os.sep}xlwt{os.sep}__init__.py", "xlwt", "Workbook") # from xlwt import Workbook
            
            filename = xls_path
            try:
                # Opening the file using 'utf-16' encoding
                file1 = io.open(filename, "r", encoding="utf-16")
                data = file1.readlines()
            except:
                # Opening the file using 'utf-8' encoding
                file1 = io.open(filename, "r", encoding="utf-8")
                data = file1.readlines()
                                
            # Creating a workbook object
            xldoc = Workbook()
            # Adding a sheet to the workbook object
            sheet = xldoc.add_sheet("Sheet1", cell_overwrite_ok=True)
            # Iterating and saving the data to sheet
            for i, row in enumerate(data):
                # Two things are done here
                # Removeing the '\n' which comes while reading the file using io.open
                # Getting the values after splitting using '\t'
                for j, val in enumerate(row.replace('\n', '').split('\t')):
                    sheet.write(i, j, val)

            # Saving the file as an excel file
            xldoc.save(xls_path)

            p.save_book_as(file_name=xls_path,
                           dest_file_name=xlsx_path)
    except Exception as e:
        print("\x1B[" + "31;40mAn error occurred\x1B[" + "0m")
        PrintException()
        raise e

if module == "getActiveCell":
    
    result = GetParams("result")

    abc = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v',
           'w', 'x', 'y', 'z']

    try:
        col = int(wb.app.selection.column)
        row = wb.app.selection.row

        length = len(abc)
        if col > length:
            excess = col // length
            mod = col % length

            col = abc[excess - 1] + abc[mod - 1]
        else:
            col = abc[col - 1]

        print(row, "******")
        ans = col + str(row)

        SetVar(result, ans)
    except Exception as e:
        PrintException()
        raise e

if module == "refreshPivot":
    sheet_ = GetParams("sheet")
    pivotTableName = GetParams("table")

    try:
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")

        wb.sheets[sheet].select()
        wb.api.ActiveSheet.PivotTables(pivotTableName).PivotCache().refresh()
    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "fitCells":
    try:
        sheet_ = GetParams("sheet")
        range_cell = GetParams("cell_range")
        fit = GetParams("fit")
        row_group = GetParams("row")
        col_group = GetParams("column")
        row_ungroup = GetParams("un_row")
        col_ungroup = GetParams("un_column")
        row_levels = GetParams("row_levels")
        col_levels = GetParams("col_levels")
        row_check = GetParams("row_check")
        column_check = GetParams("column_check")
        columnWidth = GetParams("columnWidth")
        rowHeight = GetParams("rowHeight")
        mergeCell = GetParams("mergeCell")
        unmergeCell = GetParams("unmergeCell")
        
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")

        sheet_selected = wb.sheets[sheet]

        if fit is None and not columnWidth and not rowHeight:
            fit = True
        elif fit is None:
            fit = False  
        else:
            fit = eval(fit)
        if mergeCell is not None: mergeCell = eval(mergeCell)
        if unmergeCell is not None: unmergeCell = eval(unmergeCell)
        if row_group is not None: row_group = eval(row_group)
        if col_group is not None: col_group = eval(col_group)
        if row_ungroup is not None: row_ungroup = eval(row_ungroup)
        if col_ungroup is not None: col_ungroup = eval(col_ungroup)

        if fit:
            sh = sheet_selected.autofit()
        if row_group: sheet_selected.range(range_cell).api.Rows.Group()
        if col_group: sheet_selected.range(range_cell).api.Columns.Group()

        if row_ungroup: sheet_selected.range(range_cell).api.Rows.Ungroup()
        if col_ungroup: sheet_selected.range(range_cell).api.Columns.Ungroup()
        if mergeCell: sheet_selected.range(range_cell).api.Merge()
        if unmergeCell: sheet_selected.range(range_cell).api.Unmerge()
        
        if row_levels: sheet_selected.api.Outline.ShowLevels(RowLevels=int(row_levels))
        if col_levels: sheet_selected.api.Outline.ShowLevels(RowLevels=0, ColumnLevels=int(col_levels))

        if columnWidth: sheet_selected.range(range_cell).api.ColumnWidth = columnWidth
        if rowHeight: sheet_selected.range(range_cell).api.RowHeight = rowHeight
        
    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

        #sheet.api.Rows("8:8").RowHeight = 74.25
        #sheet.api.Colums("A:D").ColumnWidth = 32.71
        
if module == "CloseExcel":
    kill_app = GetParams("kill_app")
    
    wb.app.api.DisplayAlerts = False
    
    if kill_app:
        if eval(kill_app) == True:
            wb.app.kill()
    else:
        try:
            wb.close()
        except Exception as e:
            print("\x1B[" + "31;40mError\x1B[" + "0m")
            PrintException()
            raise e

if module == "getFormula":
    
    cell = GetParams("cell")
    result = GetParams("var_")
    try:
        
        sheet = xls['sheet']
        formula = sheet.range(cell).formula
        formula = [list(i) for i in formula]
        SetVar(result, formula)
    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "AutoFilter":
    sheet_ = GetParams("sheet")
    columns = GetParams("columns")
    from time import sleep
    try:
        if len(columns) == 1:
            columns = columns + ":" + columns 
        
        if platform.system() == 'Windows':
            wb_sheets = [sh.name for sh in wb.sheets]
            sheet=None
            for s in wb_sheets:
                if s.strip() == sheet_:
                    sheet = s
                    break
            if not sheet:
                raise Exception(f"The name {sheet_} does not exist in the book")

            if wb.api.Sheets(sheet).AutoFilterMode == True:
                wb.api.Sheets(sheet).AutoFilterMode = False
                sleep(1)
            wb.sheets[sheet].api.Range(columns).AutoFilter(Field=1)
            
        else:
            rng = wb.sheets[sheet_].api.cells[columns]
            r = wb.sheets[sheet_].api.autofilter_range(rng)

    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "RemoveAutoFilter":
    sheet_ = GetParams("sheet")

    try:
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")

        if platform.system() == 'Windows':
            try:
                wb.api.Sheets(sheet).ShowAllData()
            except:
                pass
            wb.api.Sheets(sheet).AutoFilterMode = False
        else:
            try:
                wb.sheets[sheet].api.show_all_data()
            except:
                pass
            wb.sheets[sheet].api.autofilter_mode = False

    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "ClearFilter":
    sheet_ = GetParams("sheet")

    try:
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        if platform.system() == 'Windows':            
            try:
                wb.api.Sheets(sheet).ShowAllData()
            except:
                pass         
        else:
            wb.sheets[sheet].api.show_all_data()

    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e
    
# Duplicated - NOT VISIBLE
if module == "ClearFilters":
    sheet_ = GetParams("sheet") 
    try:       
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        
        if platform.system() == 'Windows':
            wb.sheets[sheet].api.ShowAllData()
        else:
            wb.sheets[sheet_].api.show_all_data()
        
    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "Filter":
    import re
    try:
        sheet_ = GetParams("sheet")
        start = GetParams("start")
        column = GetParams("column")
        data = GetParams("filter")
        filter_type = GetParams("filter_type")
        
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        
        if not filter_type:
            filter_type = "7"
        
        try:
            wb.sheets[sheet].activate()
        except:
            pass
        
        if ":" in start:
            # Range
            regex = "([a-zA-Z]+)([0-9]*):([a-zA-Z]+)([0-9]*)"
            matches = re.match(regex, start).groups()
            
            if matches[1] and matches[3]:
                # Full range
                start = matches[0] + matches[1]
                range_ = start
            else:
                # Columns range
                start = start.split(":")[0] + str(1)
                range_ = matches[0] + str(1) + ":" + matches[2] + str(1)
        else:
            # Column
            regex = "([a-zA-Z]+)([0-9]*)"
            matches = re.match(regex, start).groups()
            
            if matches[1]:
                # Cell
                start = start
                range_ = column + str(matches[1])
            else:
                # Column Letter
                start = start + str(1)
                range_ = column + str(1)
                
        if data.startswith("[") or data.startswith("("):
            data = eval(data)
        else:
            data = data.split(",")
        
        if filter_type in ["1", "2"]:
            if len(data) == 2:
                criteria1 = data[0]
                criteria2 = data[1]
            elif len(data) == 0:
                criteria1 = None
                criteria2 = None
            elif len(data) > 2:
                raise Exception("Filter 'xlOr' and 'xlAnd' must have one or two conditions. ['<>100'] & ['>=10', '<=20']")
            else:
                criteria1 = data[0]
                criteria2 = None
            # if filter_type == "7":
            #     if len(data) == 0:
            #         raise Exception("Filter 'xlFilterValues' need a list of one or more values. ['10', '20' , '30'...]")
        elif filter_type in ["8", "9"]:
            data = int(data[0]) + int(data[1]) * 256 + int(data[2]) * 256 * 256
        else:
            if len(data) == 0:
                data = None

        if platform.system() == 'Windows':
            n_start = wb.sheets[sheet].range(start).column
            n_end = wb.sheets[sheet].range(column + str(1)).column

            filter_column = n_end - n_start + 1

            if filter_type in ["1", "2"]:
                wb.sheets[sheet].api.Range(range_).AutoFilter(filter_column, Criteria1=criteria1, Criteria2=criteria2, Operator=filter_type)
            # elif filter_type == "7":
            #     wb.sheets[sheet].api.Range(range_).AutoFilter(filter_column, data, filter_type)
            else:
                wb.sheets[sheet].api.Range(range_).AutoFilter(filter_column, Criteria1=data, Operator=filter_type)
        else:
            n_start = wb.sheets[sheet].range(start).column
            n_end = wb.sheets[sheet].range(column + str(1)).column

            filter_column = n_end - n_start + 1
            
            rng = wb.sheets[sheet].api.cells[range_]
            if filter_type in ["1", "2"]:
                r = wb.sheets[sheet].api.autofilter_range(rng, field=filter_column, operator=filter_type, criteria1=criteria1, criteria2=criteria2)
            else:
                r = wb.sheets[sheet].api.autofilter_range(rng, field=filter_column, operator=filter_type, criteria1=data)
        
    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "DateFilter":

    try:
        sheet_ = GetParams("sheet")
        start = GetParams("start")
        column = GetParams("column")
        data = GetParams("filter")
        filter_type = GetParams("filter_type")
        filter_type = int(filter_type)

        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        try:
            wb.sheets[sheet].activate()
        except:
            pass
        if ":" in start:
            range_ = start
            start = start.split(":")[0]
        else:
            start = start + str(1)
            range_ = column + str(1)
        if data.startswith("[") or data.startswith("("):
            data = eval(data)
        else:
            data = data.split(",")

        criteria_list = []
        for d in data:
            criteria_list.append(int(filter_type))
            criteria_list.append(d)
        
        n_start = wb.sheets[sheet].range(start).column
        n_end = wb.sheets[sheet].range(column + str(1)).column
        filter_column = n_end - n_start + 1

        wb.sheets[sheet].api.Range(range_).AutoFilter(Field=filter_column, Operator=xw.constants.AutoFilterOperator.xlFilterValues, Criteria2=criteria_list)
        
    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "AdvancedFilter":

    try:
        sheet_ = GetParams("sheet")
        range = GetParams("range")
        filter = GetParams("filter")
        unique = GetParams("unique")
        copy = GetParams("copy")
        paste = GetParams("paste")
        
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        
        if unique:
            unique_ = eval(unique)
        else:
            unique_ = False
        
        if platform.system() == 'Windows':
            
            if filter:
                filter_ = wb.sheets[sheet].api.Range(filter)
            else:
                filter_ = None 
            
            if copy and eval(copy):
                paste_ = wb.sheets[sheet].api.Range(paste)
                wb.sheets[sheet].api.Range(range).AdvancedFilter(2, CriteriaRange=filter_, CopyToRange=paste_, Unique=unique_)
            else:
                wb.sheets[sheet].api.Range(range).AdvancedFilter(1, CriteriaRange=filter_, CopyToRange=None, Unique=unique_)
        
        else:
            
            if filter:
                filter_ = wb.sheets[sheet].api.cells[filter]
            else:
                filter_ = None 

            rng = wb.sheets[sheet].api.cells[range]
            if copy and eval(copy):
                paste_ = wb.sheets[sheet].api.cells[paste]
                wb.sheets[sheet].api.advanced_filter(rng, action=2, criteria_range=filter_, copy_to_range=paste_, unique=unique_)
            else:
                wb.sheets[sheet].api.advanced_filter(rng, action=1, criteria_range=filter_, copy_to_range=None, unique=unique_)

    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "rename_sheet":
    sheet_ = GetParams("sheet")
    name = GetParams("name")  

    try:
        
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        
        wb.sheets[sheet].select()
        wb.sheets[sheet].name = name

    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "style_cells":
    sheet_ = GetParams("sheet_name")
    range_ = GetParams("cell_range")
    position = GetParams("position")
    line_style = GetParams("lineStyle")

    font_size = GetParams("size")
    bold = GetParams("bold")
    underline = GetParams("underline")
    italic = GetParams("italic")
    adjustText = GetParams("adjustText")

    horizontal = GetParams("horizontal")
    vertical = GetParams("vertical")
    
    try:
        
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        
        rng = wb.sheets[sheet].api.Range(range_)
        if line_style:
            line_style = int(line_style)
            if position == "all":
                for i in range(7, 13):
                    rng.Borders(i).LineStyle = line_style
            elif position == "contour":
                for i in range(7, 11):
                    rng.Borders(i).LineStyle = line_style
            else:
                position = int(position)

                rng.Borders(position).LineStyle = line_style

        if font_size and font_size.isnumeric:
            rng.Font.Size = int(font_size)
            
        if underline:
            rng.Font.Underline = 2
            
        if bold is not None: bold = eval(bold)
        if bold:
            rng.Font.Bold = True
            
        if italic:
            rng.Font.Italic = True
            
        if adjustText is not None: adjustText = eval(adjustText)
        if adjustText:
            wb.sheets[sheet].range(range_).api.WrapText = True
            print("El check box esta activo")

        if horizontal:
            wb.sheets[sheet].range(range_).api.HorizontalAlignment = int(horizontal)
        
        if vertical:
            wb.sheets[sheet].range(range_).api.VerticalAlignment = int(vertical)
        
    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "Paste":

    sheet_ = GetParams("sheet_name")
    values = GetParams("values")
    cells = GetParams("cells")

    try:

        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")

        if values is not None:
            values = eval(values)
        
        try: 
            wb.sheets[sheet].select()    
        except:
            pass
        try: 
            wb.sheets[sheet].range(cells).select()
        except:
            pass
            
        try:
            if values:
                wb.sheets[sheet].range(cells).api.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False,
                                                               Transpose=False)
            else:
                if platform.system() == "Windows":
                    try:
                        wb.sheets[sheet].range(cells).paste()
                    except Exception as e:
                        wb.sheets[sheet].range(cells).api.PasteSpecial(Paste=-4104, Operation=-4142, SkipBlanks=False,
                                                                    Transpose=False)
                else:
                    wb.sheets[sheet].range(cells).paste()
        except:
            wb.sheets[sheet].api.PasteSpecial(Format="Texto Unicode", Link=False, DisplayAsIcon=False,
                                              NoHTMLFormatting=True)

    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "focus":
    try:
        from time import sleep
        from uiautomation import uiautomation as auto

        sleep(1)
        print(wb.app.impl.hwnd)
        name = f'\u202a{xls["path"].split(os.sep)[-1]}\u202c  -  Excel'
        control = auto.TitleBarControl(Name=name)
        control.SetFocus()
    except Exception as e:
        if e.text != 'Error no especificado':
            print("\x1B[" + "31;40mError\x1B[" + "0m")
            PrintException()
            raise e

if module == "remove_duplicate":
    sheet_ = GetParams("sheet_name")
    range_ = GetParams("range")
    column = GetParams("column")
    with_header = GetParams("header")
    
    from openpyxl.utils.cell import get_column_interval
    import re
    
    try:
        
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        
        sheet_selected = wb.sheets[sheet]
        sheet_selected.select()
        # Parses into list
        column = eval(column) if column.startswith("[") else [column]
        
        regex = "([a-zA-Z]+)([0-9]+):([a-zA-Z]+)([0-9]+)"
        matches = re.match(regex, range_).groups()
        cols = matches[0] + ":" + matches[2]
        # Using the columns as parameters it creates a list of the whole interval
        col_interval = get_column_interval(matches[0], matches[2])
        
        columns_ = []
        # It checks the relative position of the selected columns into the interval
        for col in column:
            columns_.append(col_interval.index(col)+1)
        
        sheet_selected.api.Range(range_).RemoveDuplicates(
            Columns=columns_, Header=int(bool(with_header)))
    
    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "save_mac":
    
    path_file = GetParams('path_file')
    args = {}
    if not path_file:
        path_file = xls["path"]
    if path_file.endswith(".xlsx"):
        args = {"FileFormat": 51, "ConflictResolution" :2}
    if path_file.endswith(".xls"):
        args = {"FileFormat": 56, "ConflictResolution" :2}
    if path_file.endswith(".csv"):
        args = {"FileFormat": 6, "ConflictResolution" :2}
    if path_file.endswith(".xlsm"):
        args = {"FileFormat": 52, "ConflictResolution" :2}
    
    wb.app.api.DisplayAlerts = False
    
    try:
        if path_file == xls["path"]:
            wb.api.Save()
        else:
            wb.api.SaveAs(path_file.replace("/", os.sep), CreateBackup=False, **args)
    except:
        wb.save(path_file)

if module == "save_mac_with_password":
    
    path_file = GetParams('path_file')
    password = GetParams('password')
    
    if not path_file:
        path_file = xls["path"]

    # First, it saves it, closes it and set the password
    # This is if we are trying to save a book that we're using it in the moment
    # Then if we are not using it, and we only wants to protect other excel with password
    # It will only set the password and saves it.
    try:
        wb.save(path_file)
        wb.app.quit()
        set_password(path_file, password)

    except Exception as e:
        try:
            set_password(path_file, password)
        except:
            print("\x1B[" + "31;40mError\x1B[" + "0m")
            PrintException()
            raise e

if module == "copyMove":

    
    sheet1 = GetParams('sheet_name1')
    sheet2 = GetParams('sheet_name2')
    book = GetParams("book")
    password = GetParams("password")
    copy_ = GetParams("copy")
    try:
        
        wb_sheets1 = [sh.name for sh in wb.sheets]
        
        for s in wb_sheets1:
            if s.strip() == sheet1:
                sheet_selected = wb.api.Sheets(s)
                break
        if not sheet_selected:
            raise Exception(f"The name {sheet1} does not exist in the book")

        if not password:
            password=None

        if book: 
            # app = excel.file_[excel.actual_id]['workbook'].app
            
            wb2 = wb.app.books.api.Open(book, False, None, None, password, password, IgnoreReadOnlyRecommended=True, CorruptLoad=0)
            wb_sheets2 = [sh.Name for sh in wb2.Sheets]
            if not sheet2:
                last = wb2.Sheets.Count
                destiny = wb2.Sheets(last)
            else:
                for s in wb_sheets2:
                    if s.strip() == sheet2:
                        destiny = wb2.api.Sheets(s)
                        break
                if not destiny:
                    raise Exception(f"The name {sheet2} does not exist in the book")
        else:
            if not sheet2:
                last = wb.api.Sheets.Count
                destiny = wb.api.Sheets(last)
            else:
                for s in wb_sheets1:
                    if s.strip() == sheet2:
                        destiny = wb.api.Sheets(s)
                        break
                if not destiny:
                    raise Exception(f"The name {sheet2} does not exist in the book")
                
            
        if copy_:
            sheet_selected.Copy(Before=destiny)
        else:
            sheet_selected.Move(Before=destiny)

        try:
            wb2.Save()
            wb2.Close()
        except:
            pass

    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "exportPDF":
    
    path_file = GetParams('path_file')
    sheet_ = GetParams('sheet')
    all_pages = GetParams('all_pages')
    option = GetParams('option')
    check_zoom = GetParams('check_zoom')
    select_zoom = GetParams('select_zoom')
    check_tall = GetParams('check_tall')
    input_tall = GetParams('input_tall')
    check_wide = GetParams('check_wide')
    input_wide = GetParams('input_wide')
    orientation = GetParams('orientation')

    if not sheet_:
        sheets = [wb.sheets.active]
    else:
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        
        sheets = [wb.sheets[sheet]]
        wb.sheets[sheet].activate()
    
    if all_pages and eval(all_pages):
        sheets = wb.sheets

    try:
        for sh in sheets:
            if option:
                if option == "all":
                    sh.autofit()

                if option == "columns":
                    sh.autofit('c')

                if option == "rows":
                    sh.autofit('r')

            if orientation:
                sh.api.PageSetup.Orientation = orientation
            else:
                sh.api.PageSetup.Orientation = 1
            
            if check_zoom:
                sh.api.PageSetup.Zoom = False
            elif select_zoom:
                sh.api.PageSetup.Zoom = eval(select_zoom)
            
            if check_tall:
                sh.api.PageSetup.FitToPagesTall = False
            elif input_tall:
                sh.api.PageSetup.FitToPagesTall = eval(input_tall)
                
            if check_wide:
                sh.api.PageSetup.FitToPagesWide = 1
            elif input_wide:
                sh.api.PageSetup.FitToPagesWide = eval(input_wide)
        if all_pages and eval(all_pages):
            wb.api.ExportAsFixedFormat(
                0, path_file.replace("/", os.sep), IncludeDocProperties=True, IgnorePrintAreas=False)
        elif len(sheets) == 1:
            wb.api.ActiveSheet.ExportAsFixedFormat(
                0, path_file.replace("/", os.sep), IncludeDocProperties=True, IgnorePrintAreas=False)

    except Exception as e:
        PrintException()
        raise e

if module == "ImportForm":
    form_path = GetParams('form_path')
    
    try:
        wb.api.VBProject.VBComponents.Import(form_path)

    except Exception as e:
        PrintException()
        raise e

if module == "GetCells":
    sheet_ = GetParams("sheet")
    range_ = GetParams("range")
    result = GetParams("var_")
    format_ = GetParams("date_format")    
    get_rows = GetParams("rows")
    extends = GetParams("more_data")
    
    try:
        extends = eval(extends)
        get_rows = eval(get_rows)
    except TypeError:
        pass
    
    try:
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")

        if platform.system() == 'Windows':
            sheet_selected_api = wb.sheets[sheet].api
            filtered_cells = sheet_selected_api.Range(range_).SpecialCells(12)
            cell_values = []
            
            if not get_rows:
                for r in filtered_cells.Areas:
                    range_cell = []
                    for ro in wb.sheets[sheet].api.Range(r.Address).Rows:
                        if isinstance(ro.Value, list) or isinstance(ro.Value, tuple):
                            cells = []
                            for cell in ro.Cells:
                                if isinstance(cell.Value, datetime.datetime):
                                    cells.append(get_date_with_format(cell.Value2, format_))
                                else:
                                    cells.append(cell.Value2)

                            range_cell.append(cells)
                        else:
                            if isinstance(ro.Value, datetime.datetime):
                                range_cell.append([get_date_with_format(ro.Value2, format_)])
                            else:
                                range_cell.append([ro.Value2]) 

                            
                    
                    if extends:
                        info = {"range": r.Address.replace("$", ""), "data": range_cell}
                        cell_values.append(info)
                    else:
                        cell_values = cell_values + \
                            range_cell if len(cell_values) > 0 else range_cell
            
            # # Range Areas
            # if not get_rows:
            #     for area in filtered_cells.Areas:
            #         range_cell = []
            #         for cells in area:
            #             if isinstance(cells.Value2, datetime.datetime):
            #                 range_cell.append(get_date_with_format(cells.Value2, format_))
            #             else:
            #                 range_cell.append(cells.Value2)
            #         if extends:
            #             info = {"range": area.Address.replace("$", ""), "data": range_cell}
            #             cell_values.append(info)
            #         else:
            #             cell_values.append(range_cell)
            
            # Rows
            if get_rows:
                rows = {}
                for area in filtered_cells.Areas:
                    range_cell = []
                    
                    for cells in area:
                        fila = 'Fila' + str(cells.Row)
                        if not rows.get(fila):
                            rows[fila] = []

                        if isinstance(cells.Value, datetime.datetime):
                            rows[fila].append(get_date_with_format(cells.Value2, format_))
                        else:
                            rows[fila].append(cells.Value2)            
                if extends:
                    cell_values = rows
                else:
                    for k, v in rows.items():
                        cell_values.append(v)          
        else:
            sh = wb.sheets[sheet]
            sh_range = sh.api.cells[range_]
            ra = sh.api.special_cells(sh_range, type = 12)
            cell_values = []
            for area in ra.areas():
                for row in area.rows():
                    cell_values += row.value()

        if result:
            SetVar(result, cell_values)

    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

def get_filtered_cells(sheet, range_, result, extends, excel, xls, wb):
    sheet_selected_api = wb.sheets[sheet].api
    filtered_cells = sheet_selected_api.Range(range_).SpecialCells(12)
    cell_values = []

    for r in filtered_cells.Address.split(","):
        range_cell = []
        for ro in wb.sheets[sheet].api.Range(r).Rows:
            if isinstance(ro.Value, list) or isinstance(ro.Value, tuple):
                cells = []
                for cell in ro.Cells:
                    if isinstance(cell.Value, datetime.datetime):
                        cells.append(get_date_with_format(cell.Value2))
                    else:
                        cells.append(cell.Value2)

                range_cell.append(cells)
            else:
                range_cell.append([ro.Value])
        try:
            extends = eval(extends)
        except TypeError:
            pass
        if extends:
            info = {"range": r.replace("$", ""), "data": range_cell}
            cell_values.append(info)
        else:
            cell_values = cell_values + \
                range_cell if len(cell_values) > 0 else range_cell


if module == "GetCountCells":
    sheet_ = GetParams("sheet")
    range_ = GetParams("range")
    result = GetParams("var_")
    
    try:
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        
        if platform.system() == 'Windows':
            sheet_selected_api = wb.sheets[sheet].api
            filtered_cells = sheet_selected_api.Range(range_).SpecialCells(12)
            count = 0
            for r in filtered_cells.Areas:
                range_cell = 0
                for ro in wb.sheets[sheet].api.Range(r.Address).Rows:
                    if isinstance(ro.Value, list) or isinstance(ro.Value, tuple):
                        for cell in ro.Cells:
                            count += 1
                    else:
                        count += 1

            # sheet_selected_api = wb.sheets[sheet].api
            # filtered_cells = sheet_selected_api.Range(range_).SpecialCells(12)
            # count = 0
            # for area in filtered_cells.Areas:
            #     for row in area.Rows:
            #         count += 1
        else:
            sh = wb.sheets[sheet_]
            sh_range = sh.api.cells[range_]
            ra = sh.api.special_cells(sh_range, type = 12)
            count = 0
            for area in ra.areas():
                for row in area.rows():
                    count += 1

        if result:
            SetVar(result, count)

    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "Replace":
    sheet_ = GetParams("sheet")
    range_ = GetParams("range")
    what = GetParams("what")
    replacement = GetParams("replace")
    
    try:
        what = eval(what)
    except:
        pass
    
    try:
        replacement = eval(replacement)
    except:
        pass
    
    try:
        wb.selection.api.Replace(what, replacement)

    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "Order":
    sheet_ = GetParams("sheet")
    range_ = GetParams("range")
    column = GetParams("column")
    order = GetParams("order")
    clean = GetParams("clean")
    try:
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        
        sheet_selected = wb.sheets[sheet]
        if order:
            order = int(order)
        else:
            order = 1
        if clean:
            sheet_selected.Sort.SortFields().Clear()
        sheet_selected.api.Range(range_).Sort(Key1=sheet_selected.api.Range(column), Order1=order, Orientation=1)
    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "OrderMultiple":
    sheet_ = GetParams("sheet")
    range_ = GetParams("range")
    headers = GetParams("headers")
    iframe = GetParams("iframe")

    try:
        if iframe != None:
            iframe = iframe if isinstance(iframe, dict) else eval(iframe)
            fields = iframe["table"]
            print(fields)
        else:
            fields = {}
    except Exception as e:
        PrintException()
        print(e)
    
    try:
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        
        sheet_selected = wb.api.Sheets(sheet)

        sheet_selected.Sort.SortFields.Clear()
            
        for field in fields:
            order = field['order']
            if order == 'Ascending':
                order = 1
            elif order == 'Descending':
                order = 2
            else:
                order = 1
            column = field['column'] + ":" + field['column']
            key = sheet_selected.Range(column)
            
            option = field.get('option', 0)
            
            if option == 'Normal':
                option = 0
            elif option == 'Text as Number':
                option = 1
            print(key, order, option)
            sheet_selected.Sort.SortFields.Add(Key=key, Order=order, DataOption=option)
        
        if headers and eval(headers):
            sheet_selected.Sort.Header = 1
        else:
            sheet_selected.Sort.Header = 0
        
        range_ = sheet_selected.Range(range_)
        sheet_selected.Sort.SetRange(range_)
        sheet_selected.Sort.Apply()
        
    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "refreshAll":

    try:
        wb.api.RefreshAll()
    except Exception as e:
        PrintException()
        raise e

if module == "find":      
    sheet_ = GetParams("sheet")
    range_ = GetParams("range")
    text = GetParams("text")
    look_in = GetParams("look_in")
    look_at = GetParams("look_at")
    match_case = GetParams("match_case")
    find_all = GetParams("find_all")
    var_ = GetParams("var_")

    try:
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        
        sheet_selected = wb.sheets[sheet]
        
        if not look_in:
            look_in = -4163 #xlValues
        else:
            look_in = eval(look_in)
        if not look_at:
            look_at = 2 #xlPart
        else:
            look_at = eval(look_at)
        if not match_case:
            match_case = False
        else:
            match_case = eval(match_case)
        
        if platform.system() == "Windows":        
            if find_all and eval(find_all):
                matches = []
                range_obj = sheet_selected.api.Range(range_)
                result = range_obj.Find(What=text, LookAt = look_at, LookIn = look_in, SearchDirection = 1, MatchCase = match_case)
                try:
                    while result is not None and result.Address not in matches:
                        matches.append(result.Address)
                        result = range_obj.FindNext(result)
                except:
                    while result is not None and result.address not in matches:
                        matches.append(result.address)
                        result = range_obj.FindNext(result)
            else:
                result = sheet_selected.api.Range(range_).Find(What=text, LookAt = look_at, LookIn = look_in, SearchDirection = 1, MatchCase = match_case)
                try:
                    matches = result.Address if result is not None else ""
                except:
                    matches = result.address if result is not None else ""
        else:
            if find_all and eval(find_all):
                matches = []
                range_obj = sheet_selected.range(range_)
                result = range_obj.api.find(what=text, look_at = look_at, look_in = look_in, search_direction = 1,match_case = match_case)
                while result is not None and result.get_address() not in matches:
                    result.select()
                    matches.append(result.get_address())
                    result = range_obj.api.find_next(after_=result)        
            else:
                result = sheet_selected.range(range_).api.find(what=text, look_at = look_at, look_in = look_in, search_direction = 1, match_case = match_case)
                matches = result.get_address() if result is not None else ""
            
       
        if var_:
            SetVar(var_, matches)
    except Exception as e:
        SetVar(var_, False)
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == 'find_pandas':
    sheet_name = GetParams("sheet")
    range_ = GetParams("range")
    columns = GetParams("columns")
    date_format = GetParams("format")
    text = GetParams("text")
    var_ = GetParams("var_")
    not_case = GetParams("not_case")
    
    import re
    import traceback
    import numpy as np
    from openpyxl.utils.cell import get_column_letter, column_index_from_string
    from dateutil.parser import parse
    from datetime import datetime
    
    if columns and columns != "":
        columns = columns.split(',')
        columns_ = []
        for col in columns:
            columns_.append(column_index_from_string(col)-1)
    else:
        columns_ = None

    try:
        if not sheet_name in [sh.name for sh in wb.sheets]:
            raise Exception(
                f"The name {sheet_name} does not exist in the book")
            
        excel_path = excel.file_[excel.actual_id]['path']
        wb.save() # It saves the workbook before it reads it to have every change made til that moment.
        
        if range_:
            regex = "([a-zA-Z]+)([0-9]+):([a-zA-Z]+)([0-9]+)"
            matches = re.match(regex, range_).groups()
            
            skip = int(matches[1])-1 if int(matches[1]) != 1 else None
            header = int(matches[1])-1 if int(matches[1]) != 1 else 0 # DF start form index 0, so it substracts one to the beginning of the range, unles it start in row one so it's 0.
            rows = int(matches[3]) - int(matches[1]) + 1 # It add one because range beginning and end point are included.
            cols = matches[0] + ":" + matches[2]

        df = pd.read_excel(excel_path, sheet_name=sheet_name, skiprows=skip, nrows=rows,  usecols=cols, header=None, parse_dates=columns_)
        
        if not_case and eval(not_case) == True:
            # If "Not case sensitive" option is selected, it sets every cell content that is string into lowercase together with the text to search.
            text = text.lower()
            for col in df.columns:
                for i in range(len(df[col])):
                    try:
                        df.iloc[i, col] = df.iloc[i, col].lower()
                    except:
                        continue
        
        # Leave this options to parse data into number types in case is needed later        
        # df.astype({0: "float", 1: str}, errors='ignore')
        # pd.to_numeric(df[1], errors='ignore')
        
        if columns_ != None and date_format != "":
            for col in columns_:
                try:
                    df[col].dt.strftime(date_format)    
                except Exception as e:
                    # print(e)
                    for i in range(len(df[col])):
                        try:
                            df.iloc[i, col] = parse(df.iloc[i, col]).strftime(date_format)
                        except:
                            continue            
        try:
            row, col = np.where(df == pd.to_numeric(text))
        except:   
            row, col = np.where(df == text)
        
        if len(col) > 0:
            row = (skip if skip != None else 0) + row[0] + 1 # If the range doesn't start at the beginning it add to the row number of de DF the skiped rows, in none are skipped, skip turns into 0.  
            col = get_column_letter(col[0] + 1)

            SetVar(var_, f"{col}{row}")
        else:
            SetVar(var_, "")
    
    except Exception as e:
        SetVar(var_, False)
        traceback.print_exc()
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "LockCells":
    sheet_ = GetParams("sheet")
    range_ = GetParams("range")
    locked = GetParams("locked")

    try:
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")

        locked = eval(locked) if locked else False

        sheet_selected = wb.sheets[sheet]
        result = sheet_selected.api.Range(range_).Locked = locked

    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == "add_chart":

    sheet_ = GetParams("sheet")
    range_ = GetParams("range")
    cell = GetParams("cell")
    type_ = GetParams("type")
    sheet_data = None
    
    if '!' in range_:
        sheet_data = range_.split('!')[0]
        range_ = range_.split('!')[1]
        
    try:
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")

        if not type_ or not bool(type_):
            raise Exception("The type of chart has not been selected")

        type_ = int(type_)
        # compatibilidad con versión anterior (antes usaba la api para windows)
        types_charts = {
            4: "line",
            5: "pie",
            51: "column_clustered",
            72: "xy_scatter_smooth",
            -4169: "xy_scatter"

        }
        if type_ in types_charts:
            type_ = types_charts[type_]

        sheet_selected = wb.sheets[sheet]
        if sheet_data:
            data_sheet = wb.sheets[sheet_data]
            cell = data_sheet.range(cell)
            range_ = data_sheet.range(range_)

            sheet_selected = wb.sheets[sheet]
            active_chart = sheet_selected.charts.add(cell.left, cell.top)
            active_chart.set_source_data(range_)
            active_chart.chart_type = type_
            
            
        else:  
            cell = sheet_selected.range(cell)
            range_ = sheet_selected.range(range_)

            active_chart = sheet_selected.charts.add(cell.left, cell.top)
            active_chart.set_source_data(range_)
            active_chart.chart_type = type_

    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e

if module == 'removePass':

    try:
        excel_file = GetParams('excel_file')
        new_excel_file = GetParams('new_excel_file')
        if new_excel_file:
            new_excel_file = new_excel_file.replace('/', '\\')
        excel_file = excel_file.replace('/', '\\')
        pass_excel = GetParams('pass_excel')

        import win32com.client

        if not new_excel_file:
            new_excel_file = excel_file

        xcl = win32com.client.Dispatch("Excel.Application")
        wb = xcl.Workbooks.Open(excel_file, False, False, None, pass_excel)
        xcl.DisplayAlerts = False
        wb.SaveAs(f'{new_excel_file}', None, '', '')

        
        excel.actual_id = excel.id_default

        excel.file_[excel.actual_id] = {}
        excel.file_[excel.actual_id]['workbook'] = xw.Book(new_excel_file)
        excel.file_[excel.actual_id]['app'] = excel.file_[
            excel.actual_id]['workbook'].app
        excel.file_[excel.actual_id]['sheet'] = excel.file_[
            excel.actual_id]['workbook'].sheets[0]
        excel.file_[excel.actual_id]['path'] = new_excel_file

        xw.books.active.close()
    except Exception as e:
        PrintException()
        raise e

if module == "insertImage":
    
    sheet_ = GetParams("sheet")
    image_path = GetParams("image_path")
    image_path = image_path.replace("/", os.sep)
    cell_position = GetParams("cell_position")
    try:

        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        
        sheet_selected = wb.sheets[sheet]
        cell = sheet_selected.range(cell_position)
        sheet_selected.pictures.add(image_path, top=cell.top, left=cell.left)
    except Exception as e:
        print("\x1B[" + "31;40mAn error occurred\x1B[" + "0m")
        PrintException()
        raise e

if module == "ExportChart":
    
    sheet_ = GetParams("sheet")
    index = GetParams("index")
    path = GetParams("path")
    try:

        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        
        sheet_selected = wb.sheets[sheet]
        chart = sheet_selected.api.ChartObjects(int(index))
        chart.Activate()
        chart = chart.Chart
        chart.Export(Filename=path, FilterName="PNG")

    except Exception as e:
        print("\x1B[" + "31;40mAn error occurred\x1B[" + "0m")
        PrintException()
        raise e
try:
    if module == "headless":
        
        path = GetParams("path")
        id_ = GetParams("id")

        app = xw.App(add_book=False, visible=False)

        if path:
            wb = app.books.open(path)
        else:
            wb = app.books.add()
            path = ""

        excel.actual_id = excel.id_default

        if id_:
            excel.actual_id = id_
        excel.file_[excel.actual_id] = {}
        excel.file_[excel.actual_id]['workbook'] = wb
        excel.file_[excel.actual_id]['app'] = excel.file_[
            excel.actual_id]['workbook'].app
        excel.file_[excel.actual_id]['sheet'] = excel.file_[
            excel.actual_id]['workbook'].sheets[0]
        excel.file_[excel.actual_id]['path'] = path


    if module == "write_cell":
        
        sheet_ = GetParams("sheet")
        range_ = GetParams("range")
        data = GetParams("data")

        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        
        data = eval(data)
        length = len(data[0])
        data_cells = []
        for row in data:

            if len(row) != length:
                raise Exception(
                    "All elements of a 2d list or tuple must be of the same length")
            row_list = [data[1] for data in row.items()]
            data_cells.append(row_list)

        sheet_selected = wb.sheets[sheet]

        sheet_selected.range(range_).value = data_cells

    if module == "copyPasteFormat":
        rango1 = GetParams("cell_range1")
        rango2 = GetParams("cell_range2")
        hoja1_ = GetParams("sheet_name1")
        hoja2_ = GetParams("sheet_name2")

        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == hoja1_:
                hoja1 = s
                break
        if not hoja1:
            raise Exception(f"The name {hoja1_} does not exist in the book")
        for s in wb_sheets:
            if s.strip() == hoja2_:
                hoja2 = s
                break
        if not hoja2:
            raise Exception(f"The name {hoja2_} does not exist in the book")
        
        my_old_value = wb.sheets[hoja2].range(rango2).options(ndim=2).value
        wb.sheets[hoja1].range(rango1).copy(wb.sheets[hoja2].range(rango2))
        wb.sheets[hoja2].range(rango2).value = my_old_value

    if module == "Opened":

        id_ = GetParams("id")
        name = GetParams("name")

        wb = xw.Book(name)
        wb.app.api.DisplayAlerts = False
        excel.actual_id = excel.id_default

        if id_:
            excel.actual_id = id_
        excel.file_[excel.actual_id] = {}
        excel.file_[excel.actual_id]['workbook'] = wb
        excel.file_[excel.actual_id]['app'] = excel.file_[
            excel.actual_id]['workbook'].app
        excel.file_[excel.actual_id]['sheet'] = excel.file_[
            excel.actual_id]['workbook'].sheets[0]
        excel.file_[excel.actual_id]['path'] = wb.fullname


    if module == "updateLinks":
        name = GetParams("name")
        new_name = GetParams("new_name")

        wb.api.ChangeLink(Name=name, NewName=new_name, Type=1)

    if module == "unlockBook":
        password = GetParams("password")
   
        wb.api.Unprotect(password)

    if module == "lockBook":
        password = GetParams("password")

        wb.api.Protect(password, True)
            
    if module == "unlockSheet":
        sheet_ = GetParams("sheet")
        password = GetParams("password")

        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        
        wb.sheets[sheet].api.Unprotect(password)

    if module == "lockSheet":
        sheet_ = GetParams("sheet")
        password = GetParams("password")

        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        
        if not password or password == "":
            wb.sheets[sheet].api.Protect()
        else:
            wb.sheets[sheet].api.Protect(password)
    
    if module == "xlsxToTxt":
        file_path_xlsx = GetParams("path_xlsx")
        file_path_txt = GetParams("path_txt")
        
        file_path_txt = file_path_txt.replace("/", os.sep)
        
        if file_path_xlsx:
            try:
                file_path_xlsx = file_path_xlsx.replace("/", os.sep)
                app = xw.App(add_book=False)
                app.api.DisplayAlerts = False
                app.api.Visible = False
                wb = app.api.Workbooks.Open(file_path_xlsx, False, None, None, None, None, IgnoreReadOnlyRecommended=True, CorruptLoad=0)          
                sleep(2)
                wb.SaveAs(file_path_txt,21)
            except Exception as e:
                raise e
        else:
            wb.api.SaveAs(file_path_txt,21)

    if module == "text2column":
        
        sheet_ = GetParams("sheet")
        range_ = GetParams("range")
        delimiter_options = GetParams("delimiter")
        other = GetParams("other")
        data_type = GetParams("data_type")

        options = {
            "Tab": False,
            "Semicolon": False,
            "Comma": False,
            "Space": False,
            "Other": False,
            "TextQualifier" : 1,
            "ConsecutiveDelimiter":False,
            "TextQualifier":2,
            "FieldInfo": None
        }

        if other:
            options["OtherChar"] = other

        if delimiter_options:
            options[delimiter_options] = True
        
        if data_type == "2":
            if "," not in other:
                separator = []
                for i in range(1, 100):
                    separator.append(str(i*int(other)))
                other = ",".join(separator)
            options["FieldInfo"] = [[int(value), 1] for value in other.split(",")]
            
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        
        try:
            # Call api directly, do not work everytime, eg. If working whit a shared folder through the cloud
            xlWorkbook = win32com.client.GetObject(wb.fullname)
            xlWorksheet = xlWorkbook.Sheets[sheet]
            xlWorksheet.Range(range_).TextToColumns(
            xlWorksheet.Range(range_),
            DataType = int(data_type),            
            TrailingMinusNumbers=True, 
            **options
            )
        except:
            # Call api through xlwings
            ws_ = wb.api.Sheets(sheet).Range(range_)
            wb.api.Sheets(sheet).Range(range_).TextToColumns(
                ws_,
                DataType = int(data_type),            
                TrailingMinusNumbers=True, 
                **options
            )

    if (module == "convertDecimalTimeToHours"):
        import math

        decimalTime = float(GetParams("decimalTime"))
        whereToStoreIn = GetParams("whereToStoreIn")

        hours = int(decimalTime * 24)
        minutes = int((decimalTime * 1440) %60)
        if ((decimalTime*86400%60%2) < 0.5):
            seconds = math.floor((decimalTime * 86400) %60)
        else:
            seconds = math.ceil((decimalTime * 86400) %60)

        hoursInString = "%02d:%02d:%02d" % (hours, minutes, seconds)

        SetVar(whereToStoreIn, hoursInString)

    if (module == "printSheet"):
        sheet_ = GetParams("sheet")

        xls = excel.file_[excel.actual_id]
        wb = xls['workbook']

        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")

        sheet_selected = wb.sheets[sheet].select()

        printSheet = wb.api.ActiveSheet.PrintOut()
    #VerticalAlignment
    if module == "formatText":
        sheet_ = GetParams("sheet")
        range_ = GetParams("cell_range")
        option_horizontal = GetParams("option_horizontal")
        option_vertical = GetParams("option_vertical")
        
        wb_sheets = [sh.name for sh in wb.sheets]
        sheet=None
        for s in wb_sheets:
            if s.strip() == sheet_:
                sheet = s
                break
        if not sheet:
            raise Exception(f"The name {sheet_} does not exist in the book")
        
        sheet_selected = wb.sheets[sheet]

        alignment_horizontal = {
            'align_to_data_type' : 1,
            'left' : -4131,
            'right' : -4152,
            'center' : -4108,}

        alignment_vertical = {
            'bottom' : -4107,
            'center' : -4108,
            'justify' : -4130,
            'top' : -4160,}
        
        if option_horizontal in alignment_horizontal:
            sheet_selected.range(range_).api.HorizontalAlignment = int(alignment_horizontal[option_horizontal])
        
        if option_vertical in alignment_vertical:
            sheet_selected.range(range_).api.VerticalAlignment = int(alignment_vertical[option_vertical])
            
    if module == "Maximize":

        if platform.system() == "Windows":
            wb.app.api.WindowState = -4137

    if module == "exportToJson":
        data =eval(GetParams("data"))
        json_path = GetParams("json_path")	

        for value in data[0]:
            if not value: raise Exception("Missing column name")

        obj_dict=dict.fromkeys(data[0])
        del data[0]
        json_result = list()
        for obj in data:
            temp =  dict(obj_dict)
            for i, key in enumerate(temp.keys()):
                temp[key] = obj[i]
            json_result.append(temp)

        with open(json_path, 'w',encoding='utf-8') as f:
            json.dump(json_result, f, indent=4)
        
except Exception as e:
    print("\x1B[" + "31;40mError\x1B[" + "0m")
    PrintException()
    raise e
